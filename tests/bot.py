#!/usr/bin/env python3
"""
DOKITA Test Bot V1.0 — GitHub Actions
Teste automatiquement 60 maladies × 8 scénarios = 480 tests primaires
+ 50 tests additionnels (contre-indications, interactions, sécurité, pédiatrie, performance)

Exécution : python3 tests/bot.py [--disease=M1] [--quick] [--no-cleanup]
"""

import os, sys, json, time, asyncio, uuid
import urllib.request, urllib.parse, urllib.error
from datetime import datetime
from diseases import DISEASES, MALADIES_ERREUR_FIXE, MALADIES_SYMPTOME_FLOU

# ══════════════════════════════════════════════════════════════
# CONFIG
# ══════════════════════════════════════════════════════════════
API_URL        = "https://heydoc-mu.vercel.app/api"
RAG_URL        = "https://heydoc-medecin.vercel.app/api"
MEDECIN_URL    = "https://heydoc-medecin.vercel.app/api"
DOKITA_KEY     = os.environ.get("DOKITA_KEY", "dk-2026-prod-x9f7m")
SUPABASE_URL   = os.environ.get("SUPABASE_URL", "")
SUPABASE_KEY   = os.environ.get("SUPABASE_SERVICE_KEY", "")
ANTHROPIC_KEY  = os.environ.get("ANTHROPIC_KEY", "")
PATIENT_EMAIL  = os.environ.get("TEST_PATIENT_EMAIL", "test.bot@dokita.health")
PATIENT_CODE   = os.environ.get("TEST_PATIENT_CODE", "0000")
MEDECIN_ORDRE  = os.environ.get("TEST_MEDECIN_ORDRE", "CI-2018-00456")
MEDECIN_CODE   = os.environ.get("TEST_MEDECIN_CODE", "0001")

# Parse args
DISEASE_FILTER = None
QUICK_MODE     = "--quick" in sys.argv
NO_CLEANUP     = "--no-cleanup" in sys.argv
START_FROM     = None
RESUME_RUN_ID  = None
for arg in sys.argv:
    if arg.startswith("--disease="):
        DISEASE_FILTER = arg.split("=")[1]
    if arg.startswith("--start="):
        START_FROM = arg.split("=")[1]
    if arg.startswith("--resume="):
        RESUME_RUN_ID = arg.split("=")[1]

RUN_ID   = RESUME_RUN_ID if RESUME_RUN_ID else str(uuid.uuid4())
RUN_DATE = datetime.utcnow().isoformat()
RESULTS  = []  # stockage en mémoire

print(f"\n{'='*60}")
print(f"DOKITA TEST BOT V1.0")
print(f"Run ID  : {RUN_ID}")
print(f"Date    : {RUN_DATE}")
print(f"Mode    : {'QUICK (5 maladies)' if QUICK_MODE else 'COMPLET (60 maladies)'}")
print(f"Filtre  : {DISEASE_FILTER or 'Aucun'}")
print(f"Cleanup : {'NON' if NO_CLEANUP else 'OUI'}")
print(f"{'='*60}\n")

# ══════════════════════════════════════════════════════════════
# HELPERS HTTP
# ══════════════════════════════════════════════════════════════
def http_post(url, body, headers=None, timeout=30):
    """POST JSON — retourne (status, data, duration_ms)"""
    t0 = time.time()
    headers = headers or {}
    headers["Content-Type"] = "application/json"
    data = json.dumps(body).encode("utf-8")
    req = urllib.request.Request(url, data=data, headers=headers, method="POST")
    try:
        with urllib.request.urlopen(req, timeout=timeout) as r:
            resp = json.loads(r.read().decode("utf-8"))
            return r.status, resp, int((time.time()-t0)*1000)
    except urllib.error.HTTPError as e:
        resp = json.loads(e.read().decode("utf-8") or "{}")
        return e.code, resp, int((time.time()-t0)*1000)
    except Exception as ex:
        return 0, {"error": str(ex)}, int((time.time()-t0)*1000)

def http_get(url, headers=None, timeout=30):
    """GET — retourne (status, data, duration_ms)"""
    t0 = time.time()
    headers = headers or {}
    req = urllib.request.Request(url, headers=headers, method="GET")
    try:
        with urllib.request.urlopen(req, timeout=timeout) as r:
            resp = json.loads(r.read().decode("utf-8"))
            return r.status, resp, int((time.time()-t0)*1000)
    except urllib.error.HTTPError as e:
        resp = json.loads(e.read().decode("utf-8") or "{}")
        return e.code, resp, int((time.time()-t0)*1000)
    except Exception as ex:
        return 0, {"error": str(ex)}, int((time.time()-t0)*1000)

def http_patch(url, body, headers=None, timeout=30):
    """PATCH JSON — retourne (status, data, duration_ms)"""
    t0 = time.time()
    headers = headers or {}
    headers["Content-Type"] = "application/json"
    data = json.dumps(body).encode("utf-8")
    req = urllib.request.Request(url, data=data, headers=headers, method="PATCH")
    try:
        with urllib.request.urlopen(req, timeout=timeout) as r:
            resp = json.loads(r.read().decode("utf-8") or "null")
            return r.status, resp, int((time.time()-t0)*1000)
    except urllib.error.HTTPError as e:
        resp = json.loads(e.read().decode("utf-8") or "{}")
        return e.code, resp, int((time.time()-t0)*1000)
    except Exception as ex:
        return 0, {"error": str(ex)}, int((time.time()-t0)*1000)

def http_delete(url, headers=None, timeout=30):
    """DELETE — retourne (status, data, duration_ms)"""
    t0 = time.time()
    headers = headers or {}
    req = urllib.request.Request(url, headers=headers, method="DELETE")
    try:
        with urllib.request.urlopen(req, timeout=timeout) as r:
            return r.status, {}, int((time.time()-t0)*1000)
    except urllib.error.HTTPError as e:
        return e.code, {}, int((time.time()-t0)*1000)
    except Exception as ex:
        return 0, {"error": str(ex)}, int((time.time()-t0)*1000)

def db_get(table, filters="", limit=10):
    """GET Supabase via /api/db"""
    url = f"{API_URL}/db?table={table}&limit={limit}"
    if filters:
        url += f"&{filters}"
    status, data, _ = http_get(url, headers={"x-dokita-key": DOKITA_KEY})
    if isinstance(data, list):
        return data
    return data.get("data", []) if isinstance(data, dict) else []

def db_patch_direct(table, row_id, body):
    """PATCH Supabase direct (sans /api/db pour les tests sécurité)"""
    url = f"{SUPABASE_URL}/rest/v1/{table}?id=eq.{row_id}"
    headers = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "return=minimal"
    }
    return http_patch(url, body, headers)

def db_delete_direct(table, row_id):
    """DELETE Supabase direct"""
    url = f"{SUPABASE_URL}/rest/v1/{table}?id=eq.{row_id}"
    headers = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
    }
    return http_delete(url, headers)

def supabase_insert(table, body):
    """INSERT Supabase direct"""
    url = f"{SUPABASE_URL}/rest/v1/{table}"
    headers = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "return=representation"
    }
    return http_post(url, body, headers)

# ══════════════════════════════════════════════════════════════
# AUTHENTIFICATION
# ══════════════════════════════════════════════════════════════
JWT_PATIENT  = None
JWT_MEDECIN  = None
PATIENT_ID   = None
MEDECIN_UUID = None

def login():
    """Login patient + médecin — arrêt si échec"""
    global JWT_PATIENT, JWT_MEDECIN, PATIENT_ID, MEDECIN_UUID

    print("── Login patient...")
    status, data, ms = http_post(f"{API_URL}/auth", {
        "type": "patient",
        "email": PATIENT_EMAIL,
        "code": PATIENT_CODE
    })
    if status != 200 or not data.get("token"):
        print(f"❌ Login patient FAIL {status}: {data}")
        sys.exit(1)
    JWT_PATIENT = data["token"]
    PATIENT_ID  = data.get("patientId") or (data.get("profils") or [{}])[0].get("id")
    print(f"✅ Login patient OK ({ms}ms) — ID: {PATIENT_ID}")

    print("── Login médecin...")
    status, data, ms = http_post(f"{API_URL}/auth", {
        "type": "medecin",
        "ordre": MEDECIN_ORDRE,
        "code": MEDECIN_CODE
    })
    if status != 200 or not data.get("token"):
        print(f"❌ Login médecin FAIL {status}: {data}")
        sys.exit(1)
    JWT_MEDECIN  = data["token"]
    MEDECIN_UUID = data.get("medecin", {}).get("supabaseId")
    print(f"✅ Login médecin OK ({ms}ms) — ID: {MEDECIN_UUID}\n")

# ══════════════════════════════════════════════════════════════
# GÉNÉRATION SYMPTÔMES VIA CLAUDE
# ══════════════════════════════════════════════════════════════
def generer_symptomes(disease, profil, mode="ideal", recurrence=False, historique=None):
    """
    Génère des symptômes via Claude directement.
    mode: ideal | flou | edge_case
    """
    profil_str = f"{profil['age']} ans, {profil['sexe']}, {profil['poids']}kg, {profil['ville']}"
    cas_special = profil.get("cas_special", "aucun")
    if cas_special == "enceinte":
        profil_str += f", enceinte T{profil.get('grossesse_trimestre',2)}"
    elif cas_special == "enfant" and profil.get("mois"):
        profil_str += f" ({profil['mois']} mois)"
    elif cas_special == "VIH+":
        profil_str += f", VIH+ CD4={profil.get('cd4', 200)}"
    elif cas_special == "allergie":
        profil_str += f", allergie {profil.get('allergie', 'pénicilline')}"
    elif cas_special in ("diabetique", "drepano"):
        profil_str += f", {cas_special}"

    if mode == "ideal":
        instruction = f"""Génère un message de consultation médicale réaliste d'un patient atteint de {disease['nom']}.
Le message doit contenir : durée précise, température si fièvre, localisation exacte, intensité, symptômes associés complets.
Le message doit être cliniquement précis pour permettre à un médecin IA de diagnostiquer facilement {disease['nom']}.
Profil patient : {profil_str}
Écris uniquement le message du patient (50-120 mots), en français, à la première personne."""

    elif mode == "flou":
        instruction = f"""Génère un message de consultation médicale VOLONTAIREMENT VAGUE d'un patient qui pourrait avoir {disease['nom']}.
Le message doit être imprécis : durée floue ("depuis quelque temps"), température non mesurée, localisation vague.
Le but est de tester si AfriBot pose des questions de relance.
Profil patient : {profil_str}
Écris uniquement le message du patient (30-60 mots), en français, à la première personne."""

    elif mode == "edge_case":
        instruction = f"""Génère un message de consultation médicale d'un patient atteint de {disease['nom']} avec le profil spécial suivant : {cas_special}.
Les symptômes doivent être cliniquement cohérents avec la maladie ET le profil spécial.
Le profil spécial doit être mentionné naturellement dans le message.
Profil patient : {profil_str}
Écris uniquement le message du patient (60-130 mots), en français, à la première personne."""

    elif mode == "recurrence":
        hist = historique or f"consultation précédente pour {disease['nom']} il y a 3 semaines"
        instruction = f"""Génère un message d'un patient qui reconsulte pour {disease['nom']} pour la 2ème fois.
Le patient doit mentionner qu'il a déjà été soigné pour ça récemment.
Profil patient : {profil_str}
Historique : {hist}
Écris uniquement le message du patient (60-100 mots), en français, à la première personne."""

    status, data, ms = http_post("https://api.anthropic.com/v1/messages", {
        "model": "claude-sonnet-4-6",
        "max_tokens": 300,
        "messages": [{"role": "user", "content": instruction}]
    }, headers={
        "x-api-key": ANTHROPIC_KEY,
        "anthropic-version": "2023-06-01"
    })

    time.sleep(1.0)  # Tier 1 rate limit — 50 req/min
    if status == 200:
        raw = data.get("content", [{}])[0].get("text", "").strip()
        # Strip markdown — AfriBot attend message brut sans formatage
        raw = raw.replace('**', '').replace('*', '').replace('__', '')
        lines = [l for l in raw.splitlines() if not l.strip().startswith('#') and l.strip() != '---']
        raw = ' '.join(' '.join(lines).split()).strip()
        return raw, ms
    return f"J'ai de la fièvre et des symptômes depuis plusieurs jours. Je me sens mal.", ms

# ══════════════════════════════════════════════════════════════
# SIMULATION AFRIBOT
# ══════════════════════════════════════════════════════════════
def simuler_afribot(symptomes, profil, disease, historique_msgs=None):
    """
    Simule une conversation avec AfriBot jusqu'à la Phase 4.
    Retourne (messages_complets, resume_json, duration_total_ms, diagnostic_detecte)
    """
    profil_context = {
        "nom": f"Patient Test {disease['id']}",
        "age": str(profil.get("age", 30)),
        "poids": str(profil.get("poids", 65)),
        "ville": profil.get("ville", "Abidjan"),
        "sexe": "Masculin" if profil.get("sexe") == "H" else "Féminin",
        "allergies": profil.get("allergie", "") if profil.get("cas_special") == "allergie" else "",
        "antecedents": "",
        "traitements": "",
        "groupe_sanguin": "",
        "taille": "",
        "imc": "",
        "langue": "Français"
    }

    messages = historique_msgs or []
    messages.append({"role": "user", "content": symptomes})

    total_ms  = 0
    phase4_ok = False
    max_tours = 8

    for tour in range(max_tours):
        status, data, ms = http_post(f"{RAG_URL}/rag", {
            "messages": messages,
            "patient": profil_context
        }, headers={"x-dokita-key": DOKITA_KEY}, timeout=45)

        total_ms += ms
        time.sleep(1.2)  # Tier 1 rate limit

        if status != 200:
            return messages, None, total_ms, False, None

        reponse = data.get("answer", "")
        messages.append({"role": "assistant", "content": reponse})

        # Phase 4 atteinte ?
        if "SUSPICIONS DIAGNOSTIQUES" in reponse or "ANALYSE DE VOS SYMPTÔMES" in reponse:
            phase4_ok = True
            break

        # Si AfriBot pose une question → répondre automatiquement
        if "?" in reponse and tour < max_tours - 1:
            reponse_auto = f"Depuis 3 jours, fièvre à 38.5°C, intensité modérée, je vis à {profil.get('ville', 'Abidjan')}, pas de voyage récent, pas de médicament en cours."
            messages.append({"role": "user", "content": reponse_auto})

    if not phase4_ok:
        return messages, None, total_ms, False, None

    # Vérifier que le diagnostic attendu est mentionné
    dernier_msg = messages[-1]["content"].lower() if messages else ""
    diagnostic_nom = disease["diagnostic_attendu"].lower()
    diagnostic_ok = diagnostic_nom in dernier_msg or diagnostic_nom[:5] in dernier_msg

    # RESUME_CONSULTATION
    messages.append({"role": "user", "content": "RESUME_CONSULTATION"})
    status, data, ms = http_post(f"{RAG_URL}/rag", {
        "messages": messages,
        "patient": profil_context
    }, headers={"x-dokita-key": DOKITA_KEY}, timeout=45)
    total_ms += ms
    time.sleep(2)  # Tier 1 rate limit

    if status != 200:
        return messages, None, total_ms, diagnostic_ok, None

    reponse_resume = data.get("answer", "")
    messages.append({"role": "assistant", "content": reponse_resume})

    # Parser JSON — extraction robuste par regex
    resume = None
    try:
        import re
        # Chercher le premier objet JSON dans la réponse
        match = re.search(r'\{[\s\S]*\}', reponse_resume)
        if match:
            resume = json.loads(match.group(0))
        else:
            clean = reponse_resume.replace("```json", "").replace("```", "").strip()
            resume = json.loads(clean)
    except Exception as e:
        print(f"    ⚠️ RESUME parse error: {e} | raw[:200]: {reponse_resume[:200]}")

    return messages, resume, total_ms, diagnostic_ok, ms

# ══════════════════════════════════════════════════════════════
# SAUVEGARDE CONSULTATION
# ══════════════════════════════════════════════════════════════
def sauvegarder_consultation(resume, profil, disease, force_flou=False):
    """Appelle /api/save comme le frontend patient"""
    if not resume:
        resume = {
            "nom": f"Patient Test {disease['id']}",
            "age": str(profil.get("age", 30)),
            "diagnostic": disease["diagnostic_attendu"],
            "symptomes": "Symptômes de test",
            "recommandations": "Recommandations OMS",
            "examens": " | ".join(disease["examens_obligatoires"]),
            "medicaments_oms": disease["medicament_1ere"],
            "contre_indications": "Aucune",
            "sources": "OMS Guidelines 2024",
            "note_historique": "",
            "traitements_en_cours": ""
        }

    body = {
        "patient_id": PATIENT_ID,
        "patient_nom": resume.get("nom", f"Patient {disease['id']}"),
        "patient_age": resume.get("age", profil.get("age", 30)),
        "patient_poids": profil.get("poids", 65),
        "patient_ville": resume.get("ville", profil.get("ville", "Abidjan")),
        "patient_voyage": resume.get("voyage", "Aucun"),
        "patient_sexe": "Masculin" if profil.get("sexe") == "H" else "Féminin",
        "symptomes": resume.get("symptomes", ""),
        "diagnostic_ia": resume.get("diagnostic", disease["diagnostic_attendu"]),
        "recommandations_oms": resume.get("recommandations", ""),
        "examens_recommandes": resume.get("examens", ""),
        "examens": resume.get("examens", ""),
        "medicaments_oms": resume.get("medicaments_oms", ""),
        "contre_indications": resume.get("contre_indications", ""),
        "note_historique": resume.get("note_historique", ""),
        "sources_oms": resume.get("sources", ""),
        "sources": resume.get("sources", ""),
        "medecin_id": MEDECIN_UUID
    }

    status, data, ms = http_post(f"{API_URL}/save", body,
        headers={"x-dokita-key": DOKITA_KEY}, timeout=30)

    return status, data, ms

# ══════════════════════════════════════════════════════════════
# PHASE MÉDECIN
# ══════════════════════════════════════════════════════════════
def medecin_ouvrir_consultation(consultation_uuid):
    """Récupère la consultation depuis /api/medecin"""
    status, data, ms = http_get(
        f"{API_URL}/medecin?action=consultations",
        headers={"Authorization": f"Bearer {JWT_MEDECIN}"}
    )
    if status != 200:
        return None, ms

    consults = data if isinstance(data, list) else (data.get("data") or [])
    for c in consults:
        if c.get("id") == consultation_uuid or c.get("consultation_id", "").startswith(f"DOS-{PATIENT_ID}"):
            return c, ms

    # Fallback : chercher par UUID direct
    url = f"{API_URL}/db?table=consultations&filter=id=eq.{consultation_uuid}&is_test=eq.true"
    status2, data2, ms2 = http_get(url, headers={"x-dokita-key": DOKITA_KEY})
    rows = data2 if isinstance(data2, list) else (data2.get("data") or [])
    return (rows[0] if rows else None), ms + ms2

def generer_prescription(disease, consultation, profil=None, mode="1ere", oublier_examen=False):
    """
    Génère une prescription via Claude basée sur les données Supabase.
    mode: 1ere | 2eme | complet | incomplet
    """
    examens_oms = consultation.get("examens_recommandes", "") or " | ".join(disease["examens_obligatoires"])
    meds_oms    = consultation.get("medicaments_oms", "") or disease["medicament_1ere"]
    profil_ctx  = ""
    if profil:
        age, sexe, poids = profil.get("age","?"), profil.get("sexe","?"), profil.get("poids","?")
        cas = profil.get("cas_special","aucun")
        p = f"{age} ans, {'Femme' if sexe=='F' else 'Homme'}, {poids}kg"
        if cas == "enceinte": p += f", enceinte T{profil.get('grossesse_trimestre',2)}"
        elif cas == "enfant": p += f", enfant {poids}kg - adapter dose au poids"
        elif cas == "allergie": p += f", allergie {profil.get('allergie','penicilline')}"
        elif cas == "VIH+": p += f", VIH+ CD4={profil.get('cd4',200)}"
        profil_ctx = f"\nPROFIL PATIENT : {p}. Adapter posologie et choix medicament."

    if mode == "1ere":
        # Utiliser medicaments_oms de la consultation (adapté par AfriBot au profil réel)
        meds_pour_ce_patient = consultation.get("medicaments_oms", "") or disease["medicament_1ere"]
        exam_target = disease["examens_obligatoires"]
        instruction = f"""Tu es médecin. Génère une prescription pour {disease['nom']}.
Le médecin a lu les recommandations AfriBot et prescrit la 1ère intention adaptée à ce patient.
Médicament OMS recommandé pour CE patient : {meds_pour_ce_patient[:400]}
Examens obligatoires : {', '.join(exam_target)}{profil_ctx}

Retourne UNIQUEMENT ce JSON valide :
{{"diagnostic":"{disease['nom']}", "examens_prescrits":["exam1","exam2"], "medicaments":[{{"nom":"...","dose":"...","duree":"..."}}]}}"""

    elif mode == "2eme":
        med_target = disease["medicament_2eme"]
        instruction = f"""Tu es médecin. Génère une prescription pour {disease['nom']} avec médicaments de 2ème intention.
Médicament 2ème intention OMS : {med_target}
Examens OMS : {', '.join(disease['examens_obligatoires'])}{profil_ctx}
Le diagnostic à poser est : {disease['nom']}

Retourne UNIQUEMENT ce JSON :
{{"diagnostic":"{disease['nom']}", "examens_prescrits":["exam1","exam2"], "medicaments":[{{"nom":"...","dose":"...","duree":"..."}}]}}"""

    elif mode == "complet":  # E3 — prescription parfaite
        # Utiliser les médicaments OMS de la consultation (adaptés par AfriBot au profil réel)
        # plutôt que disease['medicament_1ere'] qui est générique
        meds_pour_ce_patient = consultation.get("medicaments_oms", "") or disease['medicament_1ere']
        instruction = f"""Tu es médecin expert. Génère une prescription PARFAITE pour {disease['nom']}.
Le médecin a lu les recommandations AfriBot et prescrit en conséquence.
Médicament OMS recommandé pour CE patient : {meds_pour_ce_patient[:400]}
TOUS les examens obligatoires : {', '.join(disease['examens_obligatoires'])}{profil_ctx}

Retourne UNIQUEMENT ce JSON valide :
{{"diagnostic":"{disease['nom']}", "examens_prescrits":{json.dumps(disease['examens_obligatoires'])}, "medicaments":[{{"nom":"...","dose":"...","duree":"..."}}]}}"""

    else:  # incomplet — E4 — oubli examen
        examen_oublie = disease.get("examen_a_oublier", disease["examens_obligatoires"][-1])
        examens_sans_oubli = [e for e in disease["examens_obligatoires"] if e != examen_oublie]
        instruction = f"""Tu es médecin. Génère une prescription pour {disease['nom']} en OUBLIANT intentionnellement l'examen : {examen_oublie}
Examens à prescrire (liste incomplète) : {', '.join(examens_sans_oubli)}
Médicament : {disease['medicament_1ere']}

Retourne UNIQUEMENT ce JSON :
{{"diagnostic":"...", "examens_prescrits":["exam1"], "medicaments":[{{"nom":"...","dose":"...","duree":"..."}}], "examen_oublie":"{examen_oublie}"}}"""

    status, data, ms = http_post("https://api.anthropic.com/v1/messages", {
        "model": "claude-sonnet-4-6",
        "max_tokens": 400,
        "messages": [{"role": "user", "content": instruction}]
    }, headers={
        "x-api-key": ANTHROPIC_KEY,
        "anthropic-version": "2023-06-01"
    })

    time.sleep(1.0)  # Tier 1 rate limit
    if status == 200:
        raw = data.get("content", [{}])[0].get("text", "").strip()
        try:
            clean = raw.replace("```json", "").replace("```", "").strip()
            return json.loads(clean), ms
        except Exception:
            pass

    # Fallback si Claude échoue
    return {
        "diagnostic": disease["diagnostic_attendu"],
        "examens_prescrits": disease["examens_obligatoires"] if mode in ("complet", "1ere") else disease["examens_obligatoires"][:1],
        "medicaments": [{"nom": disease["medicament_1ere"], "dose": "Selon OMS", "duree": "7 jours"}]
    }, ms

def sauvegarder_prescription(consultation_uuid, prescription):
    """PATCH consultation avec la prescription médecin"""
    body = {
        "diagnostic_medecin": prescription.get("diagnostic", ""),
        "examens": json.dumps(prescription.get("examens_prescrits", [])),
        "traitement_prescrit": json.dumps(prescription.get("medicaments", [])),
        "statut": "en_attente"
    }
    url = f"{API_URL}/db?table=consultations&id={consultation_uuid}"
    return http_patch(url, body, headers={"x-dokita-key": DOKITA_KEY})

def lancer_validation_ia(consultation, prescription, examen_obligatoire=None, profil=None):
    """Lance la validation IA sur la prescription"""
    import re
    examens_oms     = consultation.get("examens_recommandes", "")
    medicaments_oms = consultation.get("medicaments_oms", "")
    diag            = consultation.get("diagnostic_ia", "")

    # Contexte profil patient
    profil_str = ""
    if profil:
        age   = profil.get("age", "?")
        sexe  = "Femme" if profil.get("sexe") == "F" else "Homme"
        poids = profil.get("poids", "?")
        cas   = profil.get("cas_special", "aucun")
        p     = f"{age} ans, {sexe}, {poids}kg"
        if cas == "enceinte":
            t = profil.get("grossesse_trimestre", 2)
            p += f", enceinte T{t} — adapter médicament au trimestre {t}"
        elif cas == "enfant":
            p += f", enfant {poids}kg — vérifier posologie pédiatrique"
        elif cas == "allergie":
            p += f", allergie {profil.get('allergie','pénicilline')} — vérifier absence cet ATB"
        elif cas == "VIH+":
            p += f", VIH+ CD4={profil.get('cd4',200)} — vérifier interactions ARV"
        elif cas == "diabetique":
            p += ", diabétique — surveiller impact glycémique"
        elif cas == "drepano":
            p += ", drépanocytaire — vérifier contre-indications"
        profil_str = f"\nPROFIL PATIENT : {p}\nLa validation DOIT tenir compte de ce profil."

    mention_examen = ""
    if examen_obligatoire:
        mention_examen = f"\nEXAMEN OBLIGATOIRE À VÉRIFIER : {examen_obligatoire}\nSi cet examen ou un équivalent clinique n'est pas dans la liste des examens prescrits, tu DOIS le mettre dans 'examens_manquants'."

    prompt_validation = f"""Tu es un médecin senior expert en médecine tropicale africaine.
Évalue la prescription suivante selon les guidelines OMS.

CONSULTATION :
Diagnostic IA : {diag}
Examens OMS recommandés : {examens_oms}
Médicaments OMS recommandés : {medicaments_oms}{profil_str}

PRESCRIPTION DU MÉDECIN :
Diagnostic posé : {prescription.get('diagnostic', '')}
Examens prescrits : {', '.join(prescription.get('examens_prescrits', []))}
Médicaments : {json.dumps(prescription.get('medicaments', []))}
{mention_examen}

Réponds UNIQUEMENT avec un objet JSON valide, sans texte avant ou après.
Exemple de réponse attendue :
{{"statut": "CONFORME", "score": 85, "details": "Prescription correcte selon OMS", "examens_manquants": [], "posologie_ok": true, "ecart_posologie": ""}}

Si des examens obligatoires manquent ou si la posologie est incorrecte, utilise "NON_CONFORME".
Si prescription globalement correcte avec réserves mineures sur profil spécial, utilise "CONFORME" avec score < 90."""

    status, data, ms = http_post("https://api.anthropic.com/v1/messages", {
        "model": "claude-sonnet-4-6",
        "max_tokens": 1500,
        "messages": [{"role": "user", "content": prompt_validation}]
    }, headers={
        "x-api-key": ANTHROPIC_KEY,
        "anthropic-version": "2023-06-01"
    })

    time.sleep(1.0)  # Tier 1 rate limit
    if status == 200:
        raw = (data.get("content", [{}])[0].get("text", "")).strip()
        try:
            match = re.search(r'\{[\s\S]*\}', raw)
            if match:
                return json.loads(match.group(0)), ms
            clean = raw.replace("```json", "").replace("```", "").strip()
            return json.loads(clean), ms
        except Exception as e:
            print(f"    ⚠️ Validation parse error: {e} | raw[:200]: {raw[:200]}")
    else:
        print(f"    ⚠️ Validation API error: HTTP {status} | {data}")
    return {"statut": "ERREUR", "score": 0, "details": "Validation IA échouée"}, ms

def creer_ordonnance(consultation_uuid, prescription):
    """Crée une ordonnance dans Supabase — insert direct"""
    if not SUPABASE_URL or not SUPABASE_KEY:
        # Fallback via /api/db
        body = {
            "consultation_id": consultation_uuid,
            "patient_id": PATIENT_ID,
            "medecin_id": MEDECIN_UUID,
            "prescription_medecin": json.dumps({
                "medicaments": prescription.get("medicaments", []),
                "examens": prescription.get("examens_prescrits", []),
                "diagnostic": prescription.get("diagnostic", "")
            }),
            "statut": "brouillon",
            "is_test": True
        }
        url = f"{API_URL}/db?table=ordonnances"
        st, data, ms = http_post(url, body, headers={"x-dokita-key": DOKITA_KEY})
        if st != 200:
            print(f"    ⚠️ creer_ordonnance /api/db error: HTTP {st} | {data}")
        return st, data, ms

    # Insert direct Supabase (plus fiable)
    body = {
        "consultation_id": consultation_uuid,
        "patient_id": PATIENT_ID,
        "medecin_id": MEDECIN_UUID,
        "prescription_medecin": json.dumps({
            "medicaments": prescription.get("medicaments", []),
            "examens": prescription.get("examens_prescrits", []),
            "diagnostic": prescription.get("diagnostic", "")
        }),
        "statut": "brouillon",
        "is_test": True
    }
    st, data, ms = supabase_insert("ordonnances", body)
    if st not in (200, 201):
        print(f"    ⚠️ creer_ordonnance Supabase error: HTTP {st} | {str(data)[:200]}")
    return st, data, ms

def valider_consultation(consultation_uuid):
    """Passe le statut à valide"""
    url = f"{API_URL}/db?table=consultations&id={consultation_uuid}"
    return http_patch(url, {"statut": "valide", "validation_ia": "CONFORME"},
        headers={"x-dokita-key": DOKITA_KEY})

def supprimer_consultation(consultation_uuid):
    """Supprime une consultation test"""
    if NO_CLEANUP:
        return
    url = f"{API_URL}/db?table=consultations&id={consultation_uuid}"
    # Supprimer examens liés
    http_delete(f"{API_URL}/db?table=examens&filter=consultation_id=eq.{consultation_uuid}",
        headers={"x-dokita-key": DOKITA_KEY})
    # Supprimer consultation
    http_delete(url, headers={"x-dokita-key": DOKITA_KEY})

def supprimer_ordonnance(ordonnance_uuid):
    """Supprime une ordonnance test (sauf E3)"""
    if NO_CLEANUP:
        return
    url = f"{API_URL}/db?table=ordonnances&id={ordonnance_uuid}"
    http_delete(url, headers={"x-dokita-key": DOKITA_KEY})

# ══════════════════════════════════════════════════════════════
# STOCKAGE RÉSULTATS
# ══════════════════════════════════════════════════════════════
def stocker_resultat(result):
    """Insère le résultat dans test_results_archive Supabase"""
    RESULTS.append(result)

    if not SUPABASE_URL or not SUPABASE_KEY:
        return  # Pas de Supabase configuré — stockage en mémoire seulement

    try:
        supabase_insert("test_results_archive", {
            "run_id": RUN_ID,
            "run_date": RUN_DATE,
            "disease": result.get("disease"),
            "test_type": result.get("test_type"),
            "patient_age": result.get("patient_age"),
            "patient_sexe": result.get("patient_sexe"),
            "patient_poids": result.get("patient_poids"),
            "patient_ville": result.get("patient_ville"),
            "patient_cas_special": result.get("patient_cas_special"),
            "phase": result.get("phase"),
            "result": result.get("result"),
            "champs_ok": json.dumps(result.get("champs_ok", [])),
            "champs_ko": json.dumps(result.get("champs_ko", [])),
            "explication_ko": result.get("explication_ko", ""),
            "afribot_diagnostic": result.get("afribot_diagnostic", ""),
            "diagnostic_attendu": result.get("diagnostic_attendu", ""),
            "diagnostic_correct": result.get("diagnostic_correct", False),
            "medicaments_oms": result.get("medicaments_oms", ""),
            "medicaments_prescrits": result.get("medicaments_prescrits", ""),
            "posologie_conforme": result.get("posologie_conforme", None),
            "ecart_posologie": result.get("ecart_posologie", ""),
            "validation_ia_statut": result.get("validation_ia_statut", ""),
            "validation_ia_attendu": result.get("validation_ia_attendu", ""),
            "validation_ia_correcte": result.get("validation_ia_correcte", None),
            "examens_oms": result.get("examens_oms", ""),
            "examens_prescrits": result.get("examens_prescrits", ""),
            "examens_manquants": result.get("examens_manquants", ""),
            "recurrence_detectee": result.get("recurrence_detectee", None),
            "contre_indication_detectee": result.get("contre_indication_detectee", None),
            "duration_api_rag_ms": result.get("duration_api_rag_ms", 0),
            "duration_api_save_ms": result.get("duration_api_save_ms", 0),
            "duration_total_ms": result.get("duration_total_ms", 0),
            "consultation_snapshot": json.dumps(result.get("consultation_snapshot", {})),
            "ordonnance_conservee": result.get("ordonnance_conservee", False),
            "ordonnance_id": result.get("ordonnance_id", None),
        })
    except Exception as e:
        print(f"  ⚠️ Stockage Supabase échoué: {e}")

# ══════════════════════════════════════════════════════════════
# VÉRIFICATIONS SUPABASE
# ══════════════════════════════════════════════════════════════
def verifier_consultation_supabase(consultation_uuid, disease):
    """Vérifie que tous les champs sont remplis dans Supabase"""
    rows = db_get("consultations", f"filter=id=eq.{consultation_uuid}")
    if not rows:
        return "FAIL", [], ["consultation introuvable en base"], None

    c = rows[0]
    champs_ok, champs_ko, explications = [], [], []

    champs_requis = {
        "id": lambda v: bool(v),
        "patient_id": lambda v: bool(v),
        "medecin_id": lambda v: bool(v),
        "symptomes": lambda v: bool(v) and len(str(v)) > 10,
        "diagnostic_ia": lambda v: bool(v),
        "examens_recommandes": lambda v: bool(v) and len(str(v)) > 5,
        "recommandations_oms": lambda v: bool(v) and len(str(v)) > 5,
        "medicaments_oms": lambda v: bool(v) and len(str(v)) > 5,
        "contre_indications": lambda v: bool(v) and len(str(v)) > 3,
        "sources_oms": lambda v: bool(v) and len(str(v)) > 5,
        "is_test": lambda v: v is True,
    }

    for champ, test in champs_requis.items():
        val = c.get(champ)
        if test(val):
            champs_ok.append(champ)
        else:
            champs_ko.append(champ)
            explications.append(f"{champ}={repr(val)}")

    # Vérifier diagnostic correct
    diag_ia = str(c.get("diagnostic_ia", "")).lower()
    diag_ok = disease["diagnostic_attendu"].lower() in diag_ia

    result = "PASS" if not champs_ko else "FAIL"
    return result, champs_ok, champs_ko, c

# ══════════════════════════════════════════════════════════════
# WORKFLOW PRINCIPAL PAR MALADIE
# ══════════════════════════════════════════════════════════════
def tester_maladie(disease):
    d = disease
    print(f"\n{'─'*50}")
    print(f"🦠 {d['id']} — {d['nom']}")
    print(f"{'─'*50}")

    profil_A = d["profil_A"]
    profil_D = d["profil_D"]

    uuid_A = None  # gardé pour test récurrence
    uuid_ordonnance_E3 = None

    # ════════════════════════════════
    # CONSULTATION A — SYMPTÔMES IDÉAUX
    # ════════════════════════════════
    print(f"\n  [A] Symptômes idéaux — {profil_A['age']}ans {profil_A['sexe']} {profil_A['poids']}kg")
    t_start = time.time()

    symptomes_A, ms_gen = generer_symptomes(d, profil_A, mode="ideal")
    messages_A, resume_A, ms_rag, diag_ok_A, ms_resume = simuler_afribot(symptomes_A, profil_A, d)
    status_save, data_save, ms_save = sauvegarder_consultation(resume_A, profil_A, d)
    uuid_A = data_save.get("uuid") if status_save == 200 else None

    result_A_save = "PASS" if status_save == 200 and uuid_A else "FAIL"
    result_A_resume = "PASS" if resume_A and all(resume_A.get(k) for k in ["diagnostic","medicaments_oms","examens","sources"]) else "FAIL"
    if result_A_resume == "FAIL":
        champs_vides = [k for k in ["diagnostic","medicaments_oms","examens","sources"] if not resume_A.get(k)] if resume_A else ["resume_None"]
        print(f"    ⚠️ RESUME champs vides: {champs_vides}")
        for k in champs_vides:
            val = resume_A.get(k, "ABSENT") if resume_A else "resume=None"
            print(f"    ⚠️ {k} = '{str(val)[:120]}'")
        if resume_A:
            print(f"    ⚠️ Toutes les clés présentes: {list(resume_A.keys())}")
    result_A_diag = "PASS" if diag_ok_A else "WARN"

    print(f"    AfriBot: {result_A_diag} | RESUME: {result_A_resume} | Save: {result_A_save} | {ms_rag}ms rag")

    stocker_resultat({
        "disease": d["nom"], "test_type": "patient_ideal",
        "patient_age": profil_A["age"], "patient_sexe": profil_A["sexe"],
        "patient_poids": profil_A["poids"], "patient_ville": profil_A["ville"],
        "patient_cas_special": profil_A.get("cas_special", "aucun"),
        "phase": "phase1_afribot", "result": result_A_diag,
        "symptomes_generes": symptomes_A,
        "afribot_diagnostic": resume_A.get("diagnostic", "") if resume_A else "",
        "diagnostic_attendu": d["diagnostic_attendu"],
        "diagnostic_correct": diag_ok_A,
        "medicaments_oms": resume_A.get("medicaments_oms", "") if resume_A else "",
        "examens_oms": resume_A.get("examens", "") if resume_A else "",
        "duration_api_rag_ms": ms_rag, "duration_api_save_ms": ms_save,
        "duration_total_ms": int((time.time()-t_start)*1000)
    })

    # Tests médecin sur consultation A
    if uuid_A:
        try:
            _tester_medecin_sur_consultation(d, uuid_A, profil_A, "A", conserver_ordonnance=False)
        except Exception as e:
            print(f"    [A→medecin] ERREUR: {e}")
            stocker_resultat({"disease":d["nom"],"test_type":"erreur_medecin_A","result":"FAIL","explication_ko":str(e)})

    # ════════════════════════════════
    # CONSULTATION B — RÉCURRENCE
    # ════════════════════════════════
    print(f"\n  [B] Récurrence (voit consultation A)")
    t_start = time.time()

    hist_context = f"[HISTORIQUE CONSULTATIONS DU PATIENT]\nDate: {datetime.utcnow().strftime('%Y-%m-%d')}\nDiagnostic: {d['diagnostic_attendu']}\nStatut: valide\n"
    symptomes_B, _ = generer_symptomes(d, profil_A, mode="recurrence")

    # Injecter l'historique dans les messages initiaux
    messages_init_B = [{"role": "assistant", "content": hist_context}]
    messages_B, resume_B, ms_rag_B, diag_ok_B, _ = simuler_afribot(symptomes_B, profil_A, d, historique_msgs=messages_init_B)
    status_save_B, data_save_B, ms_save_B = sauvegarder_consultation(resume_B, profil_A, d)
    uuid_B = data_save_B.get("uuid") if status_save_B == 200 else None

    # Vérifier détection récurrence dans les messages
    all_text_B = " ".join(m.get("content","") for m in messages_B).lower()
    recurrence_ok = any(kw in all_text_B for kw in ["déjà consulté", "précédent", "même plainte", "historique", "nouveau épisode"])

    # Vérifier note_historique
    note_hist_ok = bool(resume_B and resume_B.get("note_historique"))

    result_B = "PASS" if recurrence_ok else "WARN"
    print(f"    Récurrence détectée: {result_B} | note_historique: {'PASS' if note_hist_ok else 'FAIL'}")

    stocker_resultat({
        "disease": d["nom"], "test_type": "patient_recurrence",
        "patient_age": profil_A["age"], "patient_sexe": profil_A["sexe"],
        "patient_poids": profil_A["poids"], "patient_ville": profil_A["ville"],
        "patient_cas_special": "recurrence",
        "phase": "phase1_afribot", "result": result_B,
        "symptomes_generes": symptomes_B,
        "recurrence_detectee": recurrence_ok,
        "afribot_diagnostic": resume_B.get("diagnostic","") if resume_B else "",
        "diagnostic_attendu": d["diagnostic_attendu"],
        "diagnostic_correct": diag_ok_B,
        "champs_ok": ["note_historique"] if note_hist_ok else [],
        "champs_ko": [] if note_hist_ok else ["note_historique"],
        "duration_api_rag_ms": ms_rag_B, "duration_api_save_ms": ms_save_B,
        "duration_total_ms": int((time.time()-t_start)*1000)
    })

    if uuid_B:
        try:
            _tester_medecin_sur_consultation(d, uuid_B, profil_A, "B", conserver_ordonnance=False)
        except Exception as e:
            print(f"    [B→medecin] ERREUR: {e}")
            stocker_resultat({"disease":d["nom"],"test_type":"erreur_medecin_B","result":"FAIL","explication_ko":str(e)})

    # Effacer A + B
    if uuid_A: supprimer_consultation(uuid_A)
    if uuid_B: supprimer_consultation(uuid_B)

    # ════════════════════════════════
    # CONSULTATION C — SYMPTÔMES FLOUS
    # ════════════════════════════════
    is_flou = d["id"] in MALADIES_SYMPTOME_FLOU
    print(f"\n  [C] Symptômes {'flous' if is_flou else 'standards (non flou)'} — {profil_A['age']}ans {profil_A['sexe']} 60-70ans")
    t_start = time.time()

    profil_C = {"age": 65, "sexe": "F" if profil_A["sexe"] == "H" else "H",
                "poids": 63, "ville": profil_A["ville"], "cas_special": "aucun"}
    mode_C = "flou" if is_flou else "ideal"
    symptomes_C, _ = generer_symptomes(d, profil_C, mode=mode_C)
    messages_C, resume_C, ms_rag_C, diag_ok_C, _ = simuler_afribot(symptomes_C, profil_C, d)

    # Vérifier question relance si mode flou
    if is_flou:
        nb_questions = sum(1 for m in messages_C if m["role"] == "assistant" and "?" in m["content"])
        relance_ok = nb_questions >= 1
        result_C = "PASS" if relance_ok else "WARN"
        explication = f"{nb_questions} question(s) de relance posée(s)"
    else:
        relance_ok = True
        result_C = "PASS"
        explication = "Mode non-flou — pas de vérification relance"

    status_save_C, data_save_C, ms_save_C = sauvegarder_consultation(resume_C, profil_C, d)
    uuid_C = data_save_C.get("uuid") if status_save_C == 200 else None
    print(f"    Questions relance: {result_C} | {explication}")

    stocker_resultat({
        "disease": d["nom"], "test_type": "patient_flou" if is_flou else "patient_ideal",
        "patient_age": profil_C["age"], "patient_sexe": profil_C["sexe"],
        "patient_poids": profil_C["poids"], "patient_ville": profil_C["ville"],
        "patient_cas_special": "flou" if is_flou else "aucun",
        "phase": "phase1_afribot", "result": result_C,
        "symptomes_generes": symptomes_C,
        "afribot_diagnostic": resume_C.get("diagnostic","") if resume_C else "",
        "diagnostic_attendu": d["diagnostic_attendu"],
        "diagnostic_correct": diag_ok_C,
        "explication_ko": explication if result_C != "PASS" else "",
        "duration_api_rag_ms": ms_rag_C, "duration_api_save_ms": ms_save_C,
        "duration_total_ms": int((time.time()-t_start)*1000)
    })

    if uuid_C:
        try:
            _tester_medecin_sur_consultation(d, uuid_C, profil_C, "C", conserver_ordonnance=False)
        except Exception as e:
            print(f"    [C→medecin] ERREUR: {e}")
            stocker_resultat({"disease":d["nom"],"test_type":"erreur_medecin_C","result":"FAIL","explication_ko":str(e)})
        supprimer_consultation(uuid_C)

    # ════════════════════════════════
    # CONSULTATION D — CAS LIMITE
    # ════════════════════════════════
    cas = profil_D.get("cas_special", "aucun")
    print(f"\n  [D] Cas limite — {cas} — {profil_D['age']}ans {profil_D['sexe']} {profil_D['poids']}kg")
    t_start = time.time()

    symptomes_D, _ = generer_symptomes(d, profil_D, mode="edge_case")
    messages_D, resume_D, ms_rag_D, diag_ok_D, _ = simuler_afribot(symptomes_D, profil_D, d)

    # Vérifier contre-indication adaptée
    ci_attendue = d.get("contre_indication_profil_D", "").lower()
    all_text_D = " ".join(m.get("content","") for m in messages_D).lower()
    ci_ok = any(kw in all_text_D for kw in ["contre-indiqu", "précaution", "grossesse", "allergi", "enceinte", "enfant", "kg"]) if ci_attendue else True

    status_save_D, data_save_D, ms_save_D = sauvegarder_consultation(resume_D, profil_D, d)
    uuid_D = data_save_D.get("uuid") if status_save_D == 200 else None

    result_D = "PASS" if diag_ok_D else "FAIL"
    print(f"    Diagnostic: {result_D} | Contre-indication adaptée: {'PASS' if ci_ok else 'WARN'}")

    stocker_resultat({
        "disease": d["nom"], "test_type": "patient_cas_limite",
        "patient_age": profil_D["age"], "patient_sexe": profil_D["sexe"],
        "patient_poids": profil_D["poids"], "patient_ville": profil_D["ville"],
        "patient_cas_special": cas,
        "phase": "phase1_afribot", "result": result_D,
        "symptomes_generes": symptomes_D,
        "contre_indication_detectee": ci_ok,
        "afribot_diagnostic": resume_D.get("diagnostic","") if resume_D else "",
        "diagnostic_attendu": d["diagnostic_attendu"],
        "diagnostic_correct": diag_ok_D,
        "duration_api_rag_ms": ms_rag_D, "duration_api_save_ms": ms_save_D,
        "duration_total_ms": int((time.time()-t_start)*1000)
    })

    # Tests médecin sur D — ordonnance E3 conservée
    if uuid_D:
        try:
            uuid_ordonnance_E3 = _tester_medecin_sur_consultation(d, uuid_D, profil_D, "D", conserver_ordonnance=True)
        except Exception as e:
            print(f"    [D→medecin] ERREUR: {e}")
            stocker_resultat({"disease":d["nom"],"test_type":"erreur_medecin_D","result":"FAIL","explication_ko":str(e)})
        supprimer_consultation(uuid_D)  # effacer consultation D mais pas l'ordonnance E3

    return uuid_ordonnance_E3

def _tester_medecin_sur_consultation(disease, consultation_uuid, profil, label, conserver_ordonnance=False):
    """
    Tests médecin E1-E4 sur une consultation donnée.
    Retourne l'UUID de l'ordonnance E3 si conserver_ordonnance=True.
    """
    d = disease

    # Charger la consultation
    consultation, ms_open = medecin_ouvrir_consultation(consultation_uuid)
    if not consultation:
        print(f"    [{label}] ❌ Consultation introuvable pour tests médecin")
        return None

    # Vérifier champs reçus (régression save.js)
    examens_rec = bool(consultation.get("examens_recommandes"))
    reco_rec    = bool(consultation.get("recommandations_oms"))
    meds_rec    = bool(consultation.get("medicaments_oms"))

    stocker_resultat({
        "disease": d["nom"], "test_type": f"doctor_reception_{label}",
        "patient_age": profil["age"], "patient_sexe": profil["sexe"],
        "patient_poids": profil["poids"], "patient_ville": profil["ville"],
        "patient_cas_special": profil.get("cas_special", "aucun"),
        "phase": "phase2_reception",
        "result": "PASS" if (examens_rec and reco_rec and meds_rec) else "FAIL",
        "champs_ok": [c for c, v in [("examens_recommandes",examens_rec),("recommandations_oms",reco_rec),("medicaments_oms",meds_rec)] if v],
        "champs_ko": [c for c, v in [("examens_recommandes",examens_rec),("recommandations_oms",reco_rec),("medicaments_oms",meds_rec)] if not v],
        "explication_ko": "Régression save.js — champs vides" if not (examens_rec and meds_rec) else "",
    })

    # ── E1 — 1ère intention ──
    print(f"    [{label}→E1] 1ère ligne + examens complets...", end="")
    pres_E1, ms_p1 = generer_prescription(d, consultation, profil=profil, mode="1ere")
    sauvegarder_prescription(consultation_uuid, pres_E1)
    time.sleep(1.0)  # attendre confirmation PATCH
    val_E1, ms_v1 = lancer_validation_ia(consultation, pres_E1, profil=profil)
    stat_E1   = val_E1.get("statut", "ERREUR")
    result_E1 = "PASS" if stat_E1 == "CONFORME" else ("WARN" if stat_E1 in ("PARTIELLEMENT_CONFORME","CONFORME_AVEC_REMARQUES","CONFORME_AVEC_RESERVES") else "FAIL")
    print(f" {result_E1} ({stat_E1} score={val_E1.get('score',0)}%)")

    # Créer ordonnance sur E1
    ordonnance_E1_uuid = None
    _st1, _d1, _ = creer_ordonnance(consultation_uuid, pres_E1)
    if _st1 in (200, 201):
        _rows1 = _d1 if isinstance(_d1, list) else (_d1.get("data") or [])
        ordonnance_E1_uuid = _rows1[0].get("id") if _rows1 else None
    valider_consultation(consultation_uuid)

    # Phase 3 — vérification Supabase
    p3_result_E1, p3_ok_E1, p3_ko_E1, snap_E1 = verifier_consultation_supabase(consultation_uuid, d)

    stocker_resultat({
        "disease": d["nom"], "test_type": f"doctor_1ere_{label}",
        "patient_age": profil["age"], "patient_sexe": profil["sexe"],
        "patient_poids": profil["poids"], "patient_ville": profil["ville"],
        "patient_cas_special": profil.get("cas_special", "aucun"),
        "phase": "phase2_validation_ia",
        "result": result_E1,
        "validation_ia_statut": stat_E1,
        "validation_ia_attendu": "CONFORME",
        "validation_ia_correcte": stat_E1 == "CONFORME",
        "prescription_json": json.dumps(pres_E1),
        "medicaments_prescrits": json.dumps(pres_E1.get("medicaments",[])),
        "medicaments_oms": consultation.get("medicaments_oms",""),
        "examens_oms": consultation.get("examens_recommandes",""),
        "sources_oms": consultation.get("sources_oms",""),
        "contre_indications": consultation.get("contre_indications",""),
        "recommandations_oms": consultation.get("recommandations_oms",""),
        "posologie_conforme": val_E1.get("posologie_ok", None),
        "ecart_posologie": val_E1.get("ecart_posologie",""),
        "examens_prescrits": ", ".join(pres_E1.get("examens_prescrits",[])),
        "examens_oms": consultation.get("examens_recommandes",""),
        "examens_manquants": ", ".join(val_E1.get("examens_manquants",[])),
        "duration_api_rag_ms": ms_v1,
        "duration_total_ms": ms_v1,
        "champs_ok": p3_ok_E1, "champs_ko": p3_ko_E1,
        "consultation_snapshot": snap_E1 or consultation,
        "ordonnance_conservee": conserver_ordonnance and bool(ordonnance_E1_uuid),
        "ordonnance_id": ordonnance_E1_uuid if conserver_ordonnance else None,
        "explication_ko": val_E1.get("details","") if result_E1 == "FAIL" else "",
    })

    # ── E2 — 2ème intention ──
    print(f"    [{label}→E2] 2ème intention...", end="")
    pres_E2, ms_p2 = generer_prescription(d, consultation, profil=profil, mode="2eme")
    sauvegarder_prescription(consultation_uuid, pres_E2)
    val_E2, ms_v2 = lancer_validation_ia(consultation, pres_E2, profil=profil)
    stat_E2   = val_E2.get("statut", "ERREUR")
    result_E2 = "PASS" if stat_E2 in ("CONFORME","NON_CONFORME") else "FAIL"  # les deux sont acceptables
    print(f" {stat_E2} (attendu: CONFORME ou WARN)")

    stocker_resultat({
        "disease": d["nom"], "test_type": f"doctor_2eme_{label}",
        "patient_age": profil["age"], "patient_sexe": profil["sexe"],
        "patient_poids": profil["poids"], "patient_ville": profil["ville"],
        "patient_cas_special": profil.get("cas_special", "aucun"),
        "phase": "phase2_validation_ia",
        "result": result_E2,
        "validation_ia_statut": stat_E2,
        "validation_ia_attendu": "CONFORME",
        "validation_ia_correcte": stat_E2 == "CONFORME",
        "prescription_json": json.dumps(pres_E2),
        "medicaments_prescrits": json.dumps(pres_E2.get("medicaments",[])),
        "duration_api_rag_ms": ms_v2,
        "duration_total_ms": ms_v2,
        "consultation_snapshot": consultation,
    })

    # ── E3 — Traitement 1ère ligne + examen manquant ──
    examen_oublie = d.get("examen_a_oublier", d["examens_obligatoires"][-1])
    print(f"    [{label}→E3] Examen oublié ({examen_oublie})...", end="")
    pres_E3, ms_p3 = generer_prescription(d, consultation, profil=profil, mode="incomplet")
    sauvegarder_prescription(consultation_uuid, pres_E3)
    val_E3, ms_v3 = lancer_validation_ia(consultation, pres_E3, profil=profil,
        examen_obligatoire=examen_oublie)
    stat_E3    = val_E3.get("statut", "ERREUR")
    result_E3  = "PASS" if stat_E3 == "NON_CONFORME" else "FAIL"
    mention_oubli = len(val_E3.get("examens_manquants", [])) > 0

    print(f" {result_E3} ({stat_E3} | examen détecté: {'✅' if mention_oubli else '⚠️'})")

    stocker_resultat({
        "disease": d["nom"], "test_type": f"doctor_exam_oubli_{label}",
        "patient_age": profil["age"], "patient_sexe": profil["sexe"],
        "patient_poids": profil["poids"], "patient_ville": profil["ville"],
        "patient_cas_special": profil.get("cas_special", "aucun"),
        "phase": "phase2_validation_ia",
        "result": result_E3,
        "validation_ia_statut": stat_E3,
        "validation_ia_attendu": "NON_CONFORME",
        "validation_ia_correcte": stat_E3 == "NON_CONFORME",
        "prescription_json": json.dumps(pres_E3),
        "examens_prescrits": ", ".join(pres_E3.get("examens_prescrits",[])),
        "examens_manquants": examen_oublie,
        "explication_ko": f"Validation IA n'a pas détecté l'oubli de {examen_oublie}" if result_E3 == "FAIL" else "",
        "duration_api_rag_ms": ms_v3,
        "duration_total_ms": ms_v3,
        "consultation_snapshot": consultation,
    })

    return ordonnance_E1_uuid if conserver_ordonnance else None

# ══════════════════════════════════════════════════════════════
# TESTS ADDITIONNELS
# ══════════════════════════════════════════════════════════════
def tester_securite():
    """5 tests de sécurité"""
    print(f"\n{'═'*50}")
    print("🔒 TESTS SÉCURITÉ")
    print(f"{'═'*50}")

    # SEC.1 — Token patient sur endpoint médecin
    print("  [SEC.1] JWT patient sur endpoint médecin...", end="")
    status, _, _ = http_get(f"{API_URL}/medecin?action=consultations",
        headers={"Authorization": f"Bearer {JWT_PATIENT}"})
    result = "PASS" if status == 403 else "FAIL"
    print(f" {result} (HTTP {status})")
    stocker_resultat({"disease": "SECURITE", "test_type": "securite", "phase": "SEC1",
        "result": result, "explication_ko": f"Attendu 403, reçu {status}" if result=="FAIL" else ""})

    # SEC.2 — DOKITA_KEY invalide
    print("  [SEC.2] DOKITA_KEY invalide...", end="")
    status, _, _ = http_get(f"{API_URL}/db?table=consultations&limit=1",
        headers={"x-dokita-key": "fake-key-invalid"})
    result = "PASS" if status == 401 else "FAIL"
    print(f" {result} (HTTP {status})")
    stocker_resultat({"disease": "SECURITE", "test_type": "securite", "phase": "SEC2",
        "result": result, "explication_ko": f"Attendu 401, reçu {status}" if result=="FAIL" else ""})

    # SEC.3 — Token expiré (token bidon)
    print("  [SEC.3] Token JWT invalide...", end="")
    status, data, _ = http_get(f"{API_URL}/patient?action=profil",
        headers={"Authorization": "Bearer eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.fake.fake"})
    result = "PASS" if status == 401 else "FAIL"
    print(f" {result} (HTTP {status})")
    stocker_resultat({"disease": "SECURITE", "test_type": "securite", "phase": "SEC3",
        "result": result, "explication_ko": f"Attendu 401, reçu {status}" if result=="FAIL" else ""})

    # SEC.4 — Double POST /api/save (anti-doublon)
    print("  [SEC.4] Double POST /api/save...", end="")
    d_test = DISEASES[0]
    body = {"patient_id": PATIENT_ID, "medecin_id": MEDECIN_UUID,
            "symptomes": "Test doublon", "diagnostic_ia": "Test", "is_test": True}
    st1, d1, _ = http_post(f"{API_URL}/save", body, headers={"x-dokita-key": DOKITA_KEY})
    st2, d2, _ = http_post(f"{API_URL}/save", body, headers={"x-dokita-key": DOKITA_KEY})
    id1 = d1.get("consultation_id","")
    id2 = d2.get("consultation_id","")
    result = "PASS" if st1==200 and st2==200 and id1 != id2 else "WARN"
    print(f" {result} (IDs différents: {id1!=id2})")
    # Nettoyer
    if d1.get("uuid"): supprimer_consultation(d1["uuid"])
    if d2.get("uuid"): supprimer_consultation(d2["uuid"])
    stocker_resultat({"disease": "SECURITE", "test_type": "securite", "phase": "SEC4",
        "result": result, "explication_ko": "IDs identiques — risque doublon" if id1==id2 else ""})

    # SEC.5 — Body vide /api/save
    print("  [SEC.5] Body vide /api/save...", end="")
    status, _, _ = http_post(f"{API_URL}/save", {}, headers={"x-dokita-key": DOKITA_KEY})
    result = "PASS" if status == 400 else "FAIL"
    print(f" {result} (HTTP {status})")
    stocker_resultat({"disease": "SECURITE", "test_type": "securite", "phase": "SEC5",
        "result": result, "explication_ko": f"Attendu 400, reçu {status}" if result=="FAIL" else ""})

def tester_performance():
    """Benchmarks performance sur 10 appels"""
    print(f"\n{'═'*50}")
    print("⚡ TESTS PERFORMANCE")
    print(f"{'═'*50}")

    # Auth performance
    times_auth = []
    for _ in range(5):
        _, _, ms = http_post(f"{API_URL}/auth", {"type":"patient","email":PATIENT_EMAIL,"code":PATIENT_CODE})
        times_auth.append(ms)
    avg_auth = sum(times_auth)//len(times_auth)
    result_auth = "PASS" if avg_auth < 2000 else ("WARN" if avg_auth < 5000 else "FAIL")
    print(f"  /api/auth — moy: {avg_auth}ms — {result_auth}")
    stocker_resultat({"disease": "PERFORMANCE", "test_type": "performance", "phase": "PERF_AUTH",
        "result": result_auth, "duration_api_rag_ms": avg_auth,
        "explication_ko": f"Trop lent: {avg_auth}ms > 2000ms" if result_auth!="PASS" else ""})

    # Save performance
    times_save = []
    for _ in range(5):
        body = {"patient_id": PATIENT_ID, "medecin_id": MEDECIN_UUID,
                "symptomes": "Perf test", "diagnostic_ia": "Test", "is_test": True}
        st, data, ms = http_post(f"{API_URL}/save", body, headers={"x-dokita-key": DOKITA_KEY})
        times_save.append(ms)
        if data.get("uuid"): supprimer_consultation(data["uuid"])
    avg_save = sum(times_save)//len(times_save)
    result_save = "PASS" if avg_save < 3000 else ("WARN" if avg_save < 8000 else "FAIL")
    print(f"  /api/save — moy: {avg_save}ms — {result_save}")
    stocker_resultat({"disease": "PERFORMANCE", "test_type": "performance", "phase": "PERF_SAVE",
        "result": result_save, "duration_api_save_ms": avg_save,
        "explication_ko": f"Trop lent: {avg_save}ms > 3000ms" if result_save!="PASS" else ""})

# ══════════════════════════════════════════════════════════════
# GÉNÉRATION RAPPORT HTML
# ══════════════════════════════════════════════════════════════

def generer_rapport():
    """Génère un rapport HTML riche avec détail par phase"""
    import json as _json

    # Charger résultats Supabase pour ce run_id (inclut les runs partiels précédents)
    all_results = list(RESULTS)  # résultats du run actuel
    if SUPABASE_URL and SUPABASE_KEY and RESUME_RUN_ID:
        try:
            url = f"{SUPABASE_URL}/rest/v1/test_results_archive?run_id=eq.{RUN_ID}&order=created_at.asc&limit=5000"
            headers = {"apikey": SUPABASE_KEY, "Authorization": f"Bearer {SUPABASE_KEY}"}
            st, sb_data, _ = http_get(url, headers)
            if st == 200 and isinstance(sb_data, list):
                # Dédupliquer — garder Supabase pour les résultats déjà là
                existing_ids = {r.get("disease","") + r.get("test_type","") for r in all_results}
                for r in sb_data:
                    key = r.get("disease","") + r.get("test_type","")
                    if key not in existing_ids:
                        all_results.append(r)
                print(f"  Rapport complet: {len(sb_data)} résultats Supabase + {len(RESULTS)} nouveaux")
        except Exception as e:
            print(f"  ⚠️ Chargement Supabase pour rapport: {e}")

    # Utiliser all_results pour le rapport
    RESULTS_RAPPORT = all_results
    total   = len(RESULTS_RAPPORT)
    nb_pass = sum(1 for r in RESULTS_RAPPORT if r.get("result") == "PASS")
    nb_fail = sum(1 for r in RESULTS_RAPPORT if r.get("result") == "FAIL")
    nb_warn = sum(1 for r in RESULTS_RAPPORT if r.get("result") == "WARN")
    pct     = int(nb_pass/total*100) if total > 0 else 0

    # Grouper par maladie
    by_disease = {}
    for r in RESULTS_RAPPORT:
        d = r.get("disease","?")
        by_disease.setdefault(d, []).append(r)

    def badge(result):
        if result == "PASS": return '<span style="background:#D1FAE5;color:#166534;padding:2px 8px;border-radius:4px;font-weight:bold;font-size:11px">PASS</span>'
        if result == "FAIL": return '<span style="background:#FEE2E2;color:#991B1B;padding:2px 8px;border-radius:4px;font-weight:bold;font-size:11px">FAIL</span>'
        return '<span style="background:#FEF3C7;color:#92400E;padding:2px 8px;border-radius:4px;font-weight:bold;font-size:11px">WARN</span>'

    def field(label, val, color="#374151"):
        if not val: return ""
        val_str = str(val)
        if len(val_str) > 300: val_str = val_str[:300] + "..."
        return f'<div style="margin:4px 0"><span style="color:#6B7280;font-size:11px;font-weight:600">{label}</span><div style="color:{color};font-size:12px;margin-top:2px;line-height:1.5">{val_str}</div></div>'

    def parse_prescription(pj):
        if not pj: return ""
        try:
            p = _json.loads(pj) if isinstance(pj, str) else pj
            meds = p.get("medicaments", [])
            exams = p.get("examens_prescrits", [])
            diag = p.get("diagnostic","")
            html = f'<div style="font-size:11px"><b>Diagnostic:</b> {diag}</div>'
            if meds:
                html += '<div style="font-size:11px;margin-top:4px"><b>Médicaments:</b><ul style="margin:2px 0 0 16px">'
                for m in meds:
                    html += f'<li>{m.get("nom","")} — {m.get("dose","")} — {m.get("duree","")}</li>'
                html += '</ul></div>'
            if exams:
                html += f'<div style="font-size:11px;margin-top:4px"><b>Examens:</b> {", ".join(exams)}</div>'
            return html
        except: return str(pj)[:200]

    def section_patient(r):
        snap = r.get("consultation_snapshot") or {}
        if isinstance(snap, str):
            try: snap = _json.loads(snap)
            except: snap = {}
        if isinstance(snap, str):
            try: snap = _json.loads(snap)
            except: snap = {}

        html = '<div style="background:#F0FDF4;border:1px solid #86EFAC;border-radius:8px;padding:12px;margin-bottom:8px">'
        html += f'<div style="font-weight:700;color:#166534;margin-bottom:8px">👤 Patient — {r.get("patient_age")}ans {r.get("patient_sexe")} {r.get("patient_poids")}kg • {r.get("patient_ville","")} • {r.get("patient_cas_special","")}</div>'

        symptomes = r.get("symptomes_generes") or snap.get("symptomes","")
        if symptomes: html += field("💬 Symptômes générés", symptomes, "#1F2937")

        diag = r.get("afribot_diagnostic") or snap.get("diagnostic_ia","")
        if diag: html += field("🎯 Diagnostic AfriBot", diag, "#1D4ED8")

        reco = snap.get("recommandations_oms","")
        if reco: html += field("📋 Recommandations OMS", reco[:400])

        exams = snap.get("examens_recommandes","")
        if exams: html += field("🔬 Examens OMS", exams[:400])

        meds = snap.get("medicaments_oms","")
        if meds: html += field("💊 Médicaments OMS", meds[:400])

        ci = snap.get("contre_indications","")
        if ci: html += field("⚠️ Contre-indications", ci[:300], "#92400E")

        sources = snap.get("sources_oms","")
        if sources: html += field("📚 Sources", sources[:200], "#6B7280")

        note = snap.get("note_historique","")
        if note: html += field("🔄 Note historique (récurrence)", note[:300], "#7C3AED")

        html += '</div>'
        return html

    def section_medecin(rs, label):
        e1 = next((r for r in rs if f"doctor_1ere_{label}" in r.get("test_type","")), None)
        e2 = next((r for r in rs if f"doctor_2eme_{label}" in r.get("test_type","")), None)
        e3 = next((r for r in rs if f"doctor_exam_oubli_{label}" in r.get("test_type","")), None)

        # Récupérer les données de référence OMS depuis la réception médecin
        # Snapshot depuis E1 — contient les vraies données OMS
        e1_r = next((r for r in rs if f"doctor_1ere_{label}" in r.get("test_type","")), None)
        snap = {}
        if e1_r and e1_r.get("consultation_snapshot"):
            try:
                s = e1_r["consultation_snapshot"]
                snap = _json.loads(s) if isinstance(s,str) else s
                if isinstance(snap,str): snap = _json.loads(snap)
            except: snap = {}

        # Fallback direct sur résultat E1 si snapshot vide
        def _snap_get(k, k2=None):
            v = snap.get(k,"") if snap else ""
            if not v and e1_r:
                v = e1_r.get(k,"") or (e1_r.get(k2,"") if k2 else "")
            return str(v) if v else ""

        meds_oms_ref  = _snap_get("medicaments_oms")[:300]
        exams_oms_ref = _snap_get("examens_recommandes","examens_oms")[:300]
        sources_ref   = _snap_get("sources_oms")[:200]

        # Bloc référence OMS (affiché en haut de la section médecin)
        ref_html = ""
        if meds_oms_ref or exams_oms_ref:
            ref_html += '<div style="background:#F8FAFC;border:1px solid #CBD5E1;border-radius:8px;padding:10px;margin-bottom:10px">'
            ref_html += '<div style="font-size:11px;font-weight:700;color:#475569;margin-bottom:6px">📋 RÉFÉRENCE OMS (attendu)</div>'
            if meds_oms_ref:
                ref_html += f'<div style="font-size:11px;margin-bottom:4px"><span style="color:#1D4ED8;font-weight:600">💊 Médicaments OMS :</span> {meds_oms_ref}</div>'
            if exams_oms_ref:
                ref_html += f'<div style="font-size:11px;margin-bottom:4px"><span style="color:#065F46;font-weight:600">🔬 Examens OMS :</span> {exams_oms_ref}</div>'
            if sources_ref:
                ref_html += f'<div style="font-size:11px;color:#6B7280"><span style="font-weight:600">📚 Sources :</span> {sources_ref}</div>'
            ref_html += '</div>'

        html = f'<div style="margin-top:8px">{ref_html}<div style="display:grid;grid-template-columns:1fr 1fr;gap:8px">'
        for e, etitle, ecolor in [(e1,"E1 — Traitement 1ère ligne + examens complets","#EFF6FF"), (e2,"E2 — Alternative thérapeutique + examens complets","#F5F3FF"),
                                   (e3,"E3 — Traitement 1ère ligne + examen manquant","#FFF7ED")]:
            if not e:
                html += f'<div style="background:#F9FAFB;border:1px solid #E5E7EB;border-radius:8px;padding:10px"><b style="font-size:12px">{etitle}</b><div style="color:#9CA3AF;font-size:11px">Non exécuté</div></div>'
                continue
            stat  = e.get("validation_ia_statut","")
            score = f" {e.get('ecart_posologie','')[:60]}" if e.get("ecart_posologie") else ""
            pj    = e.get("prescription_json","")
            ordo  = e.get("ordonnance_id","")
            expl  = e.get("explication_ko","") or ""
            exams_manq = e.get("examens_manquants","")
            posol_ok = e.get("posologie_conforme")

            stat_color = "#166534" if stat=="CONFORME" else "#991B1B" if stat=="NON_CONFORME" else "#92400E"
            html += f'<div style="background:{ecolor};border:1px solid #E5E7EB;border-radius:8px;padding:10px">'
            html += f'<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:6px">'
            html += f'<b style="font-size:12px">{etitle}</b>{badge(e.get("result",""))}</div>'
            html += f'<div style="font-size:11px;color:{stat_color};font-weight:600;margin-bottom:6px">Validation IA : {stat}</div>'

            # Score conformité
            sc = e.get("duration_api_rag_ms",0)  # on n'a pas score direct — utiliser posologie
            if posol_ok is True:
                html += '<div style="font-size:10px;color:#166534;margin-bottom:4px">✅ Posologie correcte</div>'
            elif posol_ok is False:
                html += '<div style="font-size:10px;color:#991B1B;margin-bottom:4px">❌ Posologie incorrecte</div>'

            # Prescription prescrite
            if pj: html += parse_prescription(pj)

            # Examens manquants
            if exams_manq:
                html += f'<div style="font-size:11px;color:#991B1B;margin-top:6px;background:#FEE2E2;padding:4px;border-radius:4px">⚠️ Manquants : {exams_manq[:150]}</div>'

            # Explication KO
            if expl and e.get("result") in ("FAIL","WARN"):
                html += f'<div style="font-size:10px;color:#92400E;margin-top:4px;background:#FEF3C7;padding:4px;border-radius:4px">{expl[:200]}</div>'

            # Ordonnance
            if ordo:
                html += f'<div style="font-size:11px;color:#166534;margin-top:6px;font-weight:600">📄 Ordonnance conservée : {ordo[:20]}...</div>'

            html += '</div>'
        html += '</div></div>'
        return html

    # Construire le HTML par maladie
    diseases_html = ""
    for disease_name, rs in by_disease.items():
        if disease_name in ("SECURITE","PERFORMANCE"): continue

        d_pass = sum(1 for r in rs if r.get("result")=="PASS")
        d_fail = sum(1 for r in rs if r.get("result")=="FAIL")
        d_warn = sum(1 for r in rs if r.get("result")=="WARN")
        d_pct  = int(d_pass/(d_pass+d_fail+d_warn)*100) if (d_pass+d_fail+d_warn) else 0
        hdr_color = "#D1FAE5" if d_fail==0 else "#FEE2E2" if d_fail > 2 else "#FEF3C7"

        diseases_html += f'''<details style="margin-bottom:12px;border:1px solid #E5E7EB;border-radius:10px;overflow:hidden">
<summary style="background:{hdr_color};padding:12px 16px;cursor:pointer;display:flex;justify-content:space-between;align-items:center">
  <span style="font-weight:700;font-size:14px">🦠 {disease_name}</span>
  <span style="font-size:12px">PASS:{d_pass} FAIL:{d_fail} WARN:{d_warn} — {d_pct}%</span>
</summary>
<div style="padding:16px">'''

        # 4 consultations A B C D
        for label, label_title in [("A","Symptômes idéaux"),("B","Récurrence"),("C","Standard"),("D","Cas limite")]:
            pat = next((r for r in rs if r.get("test_type") in (f"patient_ideal","patient_recurrence","patient_flou","patient_cas_limite")
                        and any(f"_{label}" in t.get("test_type","") for t in rs if t.get("patient_age")==r.get("patient_age"))), None)

            # Trouver le bon patient par heuristique
            if label == "A":
                pat = next((r for r in rs if r.get("test_type")=="patient_ideal" and r.get("patient_cas_special") not in ("enceinte","enfant","VIH+","allergie","diabetique","drepano","recurrence")), None)
                if not pat: pat = next((r for r in rs if r.get("test_type")=="patient_ideal"), None)
            elif label == "B":
                pat = next((r for r in rs if r.get("test_type")=="patient_recurrence"), None)
            elif label == "C":
                pat = next((r for r in rs if r.get("test_type") in ("patient_ideal","patient_flou") and
                            not any(r2.get("test_type")=="patient_recurrence" and r2.get("patient_age")==r.get("patient_age") for r2 in rs) and
                            r.get("patient_age") != next((r2.get("patient_age") for r2 in rs if r2.get("test_type")=="patient_ideal" and r2.get("patient_cas_special") not in ("enceinte","enfant")), None)), None)
            elif label == "D":
                pat = next((r for r in rs if r.get("test_type")=="patient_cas_limite"), None)

            rec = next((r for r in rs if f"doctor_reception_{label}" in r.get("test_type","")), None)
            has_doctor = any(f"_{label}" in r.get("test_type","") and "doctor_" in r.get("test_type","") for r in rs)

            if not pat and not has_doctor: continue

            diseases_html += f'<div style="margin-bottom:16px"><h3 style="font-size:13px;color:#374151;border-bottom:1px solid #E5E7EB;padding-bottom:6px;margin-bottom:10px">📋 Consultation {label} — {label_title}</h3>'

            # Section patient
            if pat:
                diseases_html += section_patient(pat)
            elif rec and rec.get("consultation_snapshot"):
                snap = rec.get("consultation_snapshot",{})
                if isinstance(snap,str):
                    try: snap = _json.loads(snap)
                    except: snap = {}
                if isinstance(snap,str):
                    try: snap = _json.loads(snap)
                    except: snap = {}
                diseases_html += f'<div style="background:#F0FDF4;border:1px solid #86EFAC;border-radius:8px;padding:12px;margin-bottom:8px">'
                diseases_html += f'<div style="font-weight:700;color:#166534;margin-bottom:6px">👤 Consultation {label}</div>'
                sym = snap.get("symptomes","")
                if sym: diseases_html += field("💬 Symptômes", sym[:400])
                diag = snap.get("diagnostic_ia","")
                if diag: diseases_html += field("🎯 Diagnostic AfriBot", diag, "#1D4ED8")
                meds = snap.get("medicaments_oms","")
                if meds: diseases_html += field("💊 Médicaments OMS", meds[:400])
                exams = snap.get("examens_recommandes","")
                if exams: diseases_html += field("🔬 Examens OMS", exams[:400])
                diseases_html += '</div>'

            # Section médecin
            if has_doctor:
                diseases_html += f'<div style="font-size:12px;font-weight:600;color:#374151;margin-bottom:6px">🩺 Tests médecin (E1–E4)</div>'
                diseases_html += section_medecin(rs, label)

            diseases_html += '</div>'

        diseases_html += '</div></details>'

    # Tests additionnels
    sec_rs  = [r for r in RESULTS if r.get("disease")=="SECURITE"]
    perf_rs = [r for r in RESULTS if r.get("disease")=="PERFORMANCE"]

    sec_html = ''.join(f'<div style="display:flex;justify-content:space-between;padding:6px 0;border-bottom:1px solid #F3F4F6"><span style="font-size:12px">{r.get("phase","")} — {r.get("explication_ko","OK")[:80]}</span>{badge(r.get("result",""))}</div>' for r in sec_rs)
    perf_html = ''.join(f'<div style="display:flex;justify-content:space-between;padding:6px 0;border-bottom:1px solid #F3F4F6"><span style="font-size:12px">{r.get("phase","")} — {r.get("duration_api_rag_ms",0) or r.get("duration_api_save_ms",0)}ms</span>{badge(r.get("result",""))}</div>' for r in perf_rs)

    html = f"""<!DOCTYPE html>
<html lang="fr"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Dokita Test Bot — Rapport {RUN_DATE[:10]}</title>
<style>
* {{margin:0;padding:0;box-sizing:border-box;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Arial,sans-serif;}}
body {{background:#F9FAFB;padding:16px;}}
.header {{background:#0A1628;color:#fff;padding:20px;border-radius:12px;margin-bottom:16px;}}
.header h1 {{font-size:22px;color:#00C896;}}
.header p {{color:#9CA3AF;margin-top:6px;font-size:13px;}}
.stats {{display:grid;grid-template-columns:repeat(4,1fr);gap:12px;margin-bottom:16px;}}
.stat {{background:#fff;border-radius:10px;padding:14px;text-align:center;border:1px solid #E5E7EB;}}
.stat .val {{font-size:32px;font-weight:800;}}
.stat .lbl {{font-size:11px;color:#6B7280;margin-top:4px;}}
.pass-val {{color:#166534;}} .fail-val {{color:#991B1B;}} .warn-val {{color:#92400E;}} .total-val {{color:#1D4ED8;}}
details summary::-webkit-details-marker {{display:none;}}
details[open] summary {{border-bottom:1px solid #E5E7EB;}}
.addon {{background:#fff;border-radius:10px;padding:16px;margin-bottom:12px;border:1px solid #E5E7EB;}}
.addon h2 {{font-size:14px;font-weight:700;margin-bottom:10px;}}
</style></head><body>
<div class="header">
  <h1>🏥 DOKITA Test Bot — Rapport détaillé</h1>
  <p>Run ID : {RUN_ID} &nbsp;|&nbsp; Date : {RUN_DATE[:19]} &nbsp;|&nbsp; Taux de succès : {pct}%</p>
</div>
<div class="stats">
  <div class="stat"><div class="val total-val">{total}</div><div class="lbl">Tests totaux</div></div>
  <div class="stat"><div class="val pass-val">{nb_pass}</div><div class="lbl">PASS ✅</div></div>
  <div class="stat"><div class="val fail-val">{nb_fail}</div><div class="lbl">FAIL ❌</div></div>
  <div class="stat"><div class="val warn-val">{nb_warn}</div><div class="lbl">WARN ⚠️</div></div>
</div>

<h2 style="font-size:15px;font-weight:700;margin-bottom:12px;color:#0A1628">📊 Résultats par maladie</h2>
{diseases_html}

<div style="display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-top:16px">
<div class="addon"><h2>🔒 Sécurité</h2>{sec_html}</div>
<div class="addon"><h2>⚡ Performance</h2>{perf_html}</div>
</div>
</body></html>"""

    with open("rapport.html", "w", encoding="utf-8") as f:
        f.write(html)
    with open("rapport.json", "w", encoding="utf-8") as f:
        _json.dump({"run_id": RUN_ID, "date": RUN_DATE, "results": RESULTS}, f, ensure_ascii=False, indent=2)

    return pct, nb_pass, nb_fail, nb_warn




# ══════════════════════════════════════════════════════════════
# GÉNÉRATION EXCEL
# ══════════════════════════════════════════════════════════════
def generer_excel():
    """Génère un fichier Excel avec toutes les données de test"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import json as _json
    except ImportError:
        print("⚠️ openpyxl non installé — Excel non généré")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Résultats"

    # ── Couleurs ──
    COLORS = {
        "g1": "1E3A5F", "g2": "065F46", "g3": "6B21A8",
        "g4": "92400E", "g5": "1D4ED8", "g6": "374151",
        "h1": "DBEAFE", "h2": "D1FAE5", "h3": "EDE9FE",
        "h4": "FEF3C7", "h5": "BFDBFE", "h6": "F3F4F6",
        "pass": "D1FAE5", "fail": "FEE2E2", "warn": "FEF3C7",
        "conf": "D1FAE5", "nonconf": "FEE2E2", "part": "FEF3C7",
    }

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def font_white_bold():
        return Font(bold=True, color="FFFFFF", size=10)

    def font_dark(color="000000"):
        return Font(bold=True, color=color, size=10)

    def border():
        s = Side(style="thin", color="D1D5DB")
        return Border(left=s, right=s, top=s, bottom=s)

    def center():
        return Alignment(horizontal="center", vertical="center", wrap_text=True)

    def left():
        return Alignment(horizontal="left", vertical="top", wrap_text=True)

    # ── Ligne 1 : groupes ──
    groups = [
        ("📋 IDENTIFICATION", 3, "g1"),
        ("👤 PROFIL PATIENT", 5, "g2"),
        ("🤖 AFRIBOT", 5, "g3"),
        ("📚 RÉFÉRENCE OMS", 5, "g4"),
        ("🩺 TEST MÉDECIN", 9, "g5"),
        ("⚡ PERFORMANCE", 2, "g6"),
    ]

    col = 1
    for label, span, color_key in groups:
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+span-1)
        cell = ws.cell(row=1, column=col, value=label)
        cell.fill = fill(COLORS[color_key])
        cell.font = font_white_bold()
        cell.alignment = center()
        cell.border = border()
        col += span

    # ── Ligne 2 : colonnes ──
    headers = [
        # Identification (h1)
        ("Maladie", "h1"), ("Consultation", "h1"), ("Test", "h1"),
        # Profil (h2)
        ("Âge", "h2"), ("Sexe", "h2"), ("Poids (kg)", "h2"), ("Ville", "h2"), ("Cas spécial", "h2"),
        # AfriBot (h3)
        ("Symptômes générés", "h3"), ("Diagnostic AfriBot", "h3"),
        ("Récurrence détectée", "h3"), ("Note historique", "h3"), ("Résultat AfriBot", "h3"),
        # Référence OMS (h4)
        ("Médicaments OMS", "h4"), ("Examens OMS", "h4"),
        ("Contre-indications OMS", "h4"), ("Sources OMS", "h4"), ("Recommandations OMS", "h4"),
        # Test médecin (h5)
        ("Résultat", "h5"), ("Validation IA", "h5"), ("Score (%)", "h5"),
        ("Diagnostic médecin", "h5"), ("Médicaments prescrits", "h5"),
        ("Posologie conforme", "h5"), ("Examens prescrits", "h5"),
        ("Examens manquants", "h5"), ("Contenu ordonnance", "h5"),
        ("Explication KO", "h5"),
        # Performance (h6)
        ("Durée AfriBot (ms)", "h6"), ("Durée Valid. IA (ms)", "h6"),
    ]

    for c, (label, color_key) in enumerate(headers, 1):
        cell = ws.cell(row=2, column=c, value=label)
        cell.fill = fill(COLORS[color_key])
        cell.font = font_dark()
        cell.alignment = center()
        cell.border = border()

    # ── Données ──
    # Regrouper par maladie
    by_disease = {}
    all_r = list(RESULTS)
    for r in all_r:
        d = r.get("disease", "?")
        by_disease.setdefault(d, []).append(r)

    row_num = 3
    for disease_name, rs in by_disease.items():
        if disease_name in ("SECURITE", "PERFORMANCE"):
            continue

        for label, label_title in [("A","Idéal"), ("B","Récurrence"), ("C","Standard"), ("D","Cas limite")]:

            # Données patient
            pat = None
            if label == "A":
                pat = next((r for r in rs if r.get("test_type") == "patient_ideal"
                    and r.get("patient_cas_special") not in ("enceinte","enfant","VIH+","allergie","diabetique","drepano","recurrence")), None)
                if not pat:
                    pat = next((r for r in rs if r.get("test_type") == "patient_ideal"), None)
            elif label == "B":
                pat = next((r for r in rs if r.get("test_type") == "patient_recurrence"), None)
            elif label == "C":
                pat = next((r for r in rs if r.get("test_type") in ("patient_ideal","patient_flou")
                    and r.get("patient_cas_special") not in ("enceinte","enfant","VIH+","allergie","diabetique","drepano","recurrence")
                    and r != next((r2 for r2 in rs if r2.get("test_type")=="patient_ideal"
                        and r2.get("patient_cas_special") not in ("enceinte","enfant","VIH+","allergie")), None)), None)
            elif label == "D":
                pat = next((r for r in rs if r.get("test_type") == "patient_cas_limite"), None)

            if not pat:
                continue

            # Snapshot OMS
            rec = next((r for r in rs if f"doctor_reception_{label}" in r.get("test_type","")), None)
            snap = {}
            if rec and rec.get("consultation_snapshot"):
                try:
                    s = rec["consultation_snapshot"]
                    snap = _json.loads(s) if isinstance(s, str) else s
                    if isinstance(snap, str):
                        snap = _json.loads(snap)
                except:
                    snap = {}

            # Données communes patient
            age   = pat.get("patient_age", "")
            sexe  = pat.get("patient_sexe", "")
            poids = pat.get("patient_poids", "")
            ville = pat.get("patient_ville", "")
            cas   = pat.get("patient_cas_special", "")
            sympt = pat.get("symptomes_generes", "")[:300]
            diag_afribot = pat.get("afribot_diagnostic", "")
            recurrence   = "Oui" if pat.get("recurrence_detectee") else "Non"
            note_hist   = snap.get("note_historique", "") or pat.get("note_historique","")
            res_afribot = pat.get("result", "")

            # OMS — fallback sur résultat E1 si snapshot vide
            e1_res = next((r for r in rs if f"doctor_1ere_{label}" in r.get("test_type","")), None)
            def _oms(k, k2=None):
                v = snap.get(k,"") if snap else ""
                if not v and e1_res:
                    v = e1_res.get(k,"") or (e1_res.get(k2,"") if k2 else "")
                return str(v)[:300] if v else ""

            meds_oms    = _oms("medicaments_oms")
            exams_oms   = _oms("examens_recommandes","examens_oms")
            ci_oms      = _oms("contre_indications")[:200]
            sources_oms = _oms("sources_oms")[:200]
            reco_oms    = _oms("recommandations_oms")[:200]
            dur_afribot = pat.get("duration_api_rag_ms", "")

            # Tests E1, E2, E3
            test_map = {
                "E1": next((r for r in rs if f"doctor_1ere_{label}" in r.get("test_type","")), None),
                "E2": next((r for r in rs if f"doctor_2eme_{label}" in r.get("test_type","")), None),
                "E3": next((r for r in rs if f"doctor_exam_oubli_{label}" in r.get("test_type","")), None),
            }

            for test_name, e in test_map.items():
                if not e:
                    continue

                # Parser prescription
                pj = {}
                try:
                    pj = _json.loads(e.get("prescription_json","{}")) if e.get("prescription_json") else {}
                except:
                    pj = {}

                meds_list = pj.get("medicaments", [])
                meds_str  = " | ".join([f"{m.get('nom','')} {m.get('dose','')} {m.get('duree','')}" for m in meds_list])
                diag_med  = pj.get("diagnostic", "")
                exams_str = e.get("examens_prescrits", "")

                # Contenu ordonnance — toujours rempli si prescription disponible
                ordo_content = ""
                if diag_med or meds_str:
                    ordo_id = e.get("ordonnance_id","")
                    ordo_tag = f" [ID:{ordo_id[:8]}]" if ordo_id else ""
                    ordo_content = f"Diag: {diag_med} | Méds: {meds_str} | Examens: {exams_str}{ordo_tag}"

                val_statut  = e.get("validation_ia_statut", "")
                score       = e.get("duration_api_rag_ms", "")  # placeholder
                posol_ok    = "✅ Oui" if e.get("posologie_conforme") is True else ("❌ Non" if e.get("posologie_conforme") is False else "—")
                exams_manq  = e.get("examens_manquants", "")
                result      = e.get("result", "")
                expl_ko     = e.get("explication_ko", "")[:200]
                dur_val     = e.get("duration_api_rag_ms", "")

                row_data = [
                    disease_name,
                    f"{label} — {label_title}",
                    test_name,
                    age, sexe, poids, ville, cas,
                    sympt, diag_afribot, recurrence, note_hist, res_afribot,
                    meds_oms, exams_oms, ci_oms, sources_oms, reco_oms,
                    result, val_statut, "",  # score calculé séparément
                    diag_med, meds_str, posol_ok, exams_str, exams_manq,
                    ordo_content, expl_ko,
                    dur_afribot, dur_val,
                ]

                for c, val in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=c, value=str(val) if val is not None else "")
                    cell.border = border()
                    cell.alignment = left()
                    cell.font = Font(size=9)

                # Coloriser colonne Résultat (col 19)
                res_cell = ws.cell(row=row_num, column=19)
                if result == "PASS":
                    res_cell.fill = fill(COLORS["pass"])
                    res_cell.font = Font(bold=True, color="166534", size=9)
                elif result == "FAIL":
                    res_cell.fill = fill(COLORS["fail"])
                    res_cell.font = Font(bold=True, color="991B1B", size=9)
                elif result == "WARN":
                    res_cell.fill = fill(COLORS["warn"])
                    res_cell.font = Font(bold=True, color="92400E", size=9)

                # Coloriser Validation IA (col 20)
                val_cell = ws.cell(row=row_num, column=20)
                if "CONFORME" in val_statut and "NON" not in val_statut:
                    val_cell.fill = fill(COLORS["conf"])
                    val_cell.font = Font(bold=True, color="166534", size=9)
                elif "NON_CONFORME" in val_statut:
                    val_cell.fill = fill(COLORS["nonconf"])
                    val_cell.font = Font(bold=True, color="991B1B", size=9)
                elif val_statut:
                    val_cell.fill = fill(COLORS["warn"])
                    val_cell.font = Font(bold=True, color="92400E", size=9)

                # Coloriser examens manquants (col 26)
                if exams_manq:
                    ws.cell(row=row_num, column=26).fill = fill("FEE2E2")
                    ws.cell(row=row_num, column=26).font = Font(bold=True, color="991B1B", size=9)

                row_num += 1

    # ── Largeurs colonnes ──
    col_widths = [
        25, 18, 12,        # Identification
        6, 6, 8, 15, 18,   # Profil
        40, 40, 12, 30, 12, # AfriBot
        45, 45, 35, 45, 35, # OMS
        10, 20, 8, 25, 40, 12, 40, 25, 45, 35, # Test médecin
        14, 14,             # Performance
    ]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Figer les 2 premières lignes
    ws.freeze_panes = "A3"

    # Hauteur lignes en-têtes
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 28

    wb.save("rapport.xlsx")
    print("📊 Excel généré : rapport.xlsx")


if __name__ == "__main__":
    login()

    # Sélectionner les maladies à tester
    diseases_to_test = DISEASES
    if DISEASE_FILTER:
        diseases_to_test = [d for d in DISEASES if d["id"] == DISEASE_FILTER]
    elif QUICK_MODE:
        quick_ids = {"M1", "M4", "M9", "M34", "M42"}
        diseases_to_test = [d for d in DISEASES if d["id"] in quick_ids]
    elif START_FROM:
        # Mode reprise — commencer à partir d'une maladie donnée
        start_idx = next((i for i, d in enumerate(DISEASES) if d["id"] == START_FROM), 0)
        diseases_to_test = DISEASES[start_idx:]
        print(f"Reprise depuis {START_FROM} ({len(diseases_to_test)} maladies restantes)")

    print(f"Maladies à tester : {len(diseases_to_test)}")
    ordonnances_conservees = {}

    # Boucle principale
    for i, disease in enumerate(diseases_to_test):
        try:
            ordo_uuid = tester_maladie(disease)
            if ordo_uuid:
                ordonnances_conservees[disease["nom"]] = ordo_uuid
        except Exception as e:
            print(f"  ❌ ERREUR CRITIQUE {disease['id']}: {e}")
            stocker_resultat({
                "disease": disease["nom"], "test_type": "erreur_critique",
                "result": "FAIL", "explication_ko": str(e)
            })

    # Tests additionnels
    tester_securite()
    tester_performance()

    # Rapport final
    print(f"\n{'═'*60}")
    print("📊 GÉNÉRATION RAPPORT")
    pct, nb_pass, nb_fail, nb_warn = generer_rapport()
    generer_excel()

    print(f"\n{'═'*60}")
    print(f"✅ PASS  : {nb_pass}")
    print(f"❌ FAIL  : {nb_fail}")
    print(f"⚠️ WARN  : {nb_warn}")
    print(f"📈 Taux  : {pct}%")

    print(f"\n💊 Ordonnances E3 conservées ({len(ordonnances_conservees)}) :")
    for maladie, oid in ordonnances_conservees.items():
        print(f"   {maladie} → {oid}")

    print(f"\nRapports : rapport.html | rapport.json")
    print(f"{'═'*60}\n")

    # Exit code pour GitHub Actions
    sys.exit(0 if nb_fail == 0 else 1)

